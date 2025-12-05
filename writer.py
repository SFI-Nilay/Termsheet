import os
from openpyxl import Workbook, load_workbook
from typing import Dict

from config import EXCEL_FILE

# -----------------------------------------------------
# Initialize Workbook
# -----------------------------------------------------
def _init_workbook() -> Workbook:
    """Initialize the workbook with a single sheet 'EXPORT' and required headers."""
    if os.path.exists(EXCEL_FILE):
        return load_workbook(EXCEL_FILE)

    wb = Workbook()
    ws = wb.active
    ws.title = "EXPORT"
    ws.append([
        "ISIN", "Bond Type", "Issuer", "Bond Size", "Currency", "Coupon",
        "Issuance Date", "Maturity Date", "Exchange Listing", "Paying Agent",
        "Moody's", "S&P", "Fitch", "Status of Notes", "Method of Distribution",
        "Syndicate", "Source", "Comment", "Date"
    ])
    wb.save(EXCEL_FILE)
    return wb

# -----------------------------------------------------
# Main Write Function
# -----------------------------------------------------
def write_to_excel(json_data: Dict) -> None:
    """
    Write parsed Term Sheet JSON data into the EXPORT sheet.
    json_data: dictionary output from parser.py
    """
    wb = _init_workbook()
    ws = wb["EXPORT"]

    row = [
        json_data.get("ISIN", ""),
        json_data.get("Bond Type", ""),
        json_data.get("Issuer", ""),
        json_data.get("Bond Size", ""),
        json_data.get("Currency", ""),
        json_data.get("Coupon", ""),
        json_data.get("Issuance Date", ""),
        json_data.get("Maturity Date", ""),
        json_data.get("Exchange Listing", ""),
        json_data.get("Paying Agent", ""),
        json_data.get("Moody's", ""),
        json_data.get("S&P", ""),
        json_data.get("Fitch", ""),
        json_data.get("Status of Notes", ""),
        json_data.get("Method of Distribution", ""),
        json_data.get("Syndicate", ""),
        json_data.get("Source", ""),
        json_data.get("Comment", ""),
        json_data.get("Date", "")
    ]

    ws.append(row)
    wb.save(EXCEL_FILE)
    print(f"✅ Data written successfully for ISIN: {json_data.get('ISIN', '')}")
